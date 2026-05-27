"""
Extract CBO Feb 2026 baseline projections to a single CSV consumed by App2.jsx
(TaxPage page 13 + EconomicImpactPage page 14).

Source: https://www.cbo.gov/publication/61882
File:   51118-2026-02-Budget-Projections-1.xlsx  (download manually from CBO)

Reads:
  Table 1-1  CBO's Baseline Budget Projections (revenue, outlays, deficits, debt, GDP)
  Table 1-3  CBO's Baseline Projections of Federal Debt (avg interest rate)
  Table 3-1  Spending Projections by Category (SS, Medicare, Medicaid, Defense, etc.)

Writes:  public/cbo_baseline_2026.csv  (long format — series, kind, 2026..2034)

Re-run whenever CBO publishes a new baseline. Point INPUT_XLSX at the new file.

Usage:
  python3 datapipeline/extract_cbo_baseline.py
"""

import csv
import pathlib
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

INPUT_XLSX = pathlib.Path.home() / "Downloads" / "51118-2026-02-Budget-Projections-1.xlsx"
OUTPUT_CSV = pathlib.Path(__file__).resolve().parent.parent / "public" / "cbo_baseline_2026.csv"

YEARS = list(range(2026, 2035))  # 9-year window 2026-2034

# (sheet, row, label, kind) — column offsets discovered via manual inspection.
# Column 2 = 2025 actual; columns 3..11 = 2026..2034.
ROWS = [
    # Spending by category (Table 3-1)
    ("Table 3-1", 11, "Social Security",       "spending"),
    ("Table 3-1", 12, "Medicare",              "spending"),
    ("Table 3-1", 13, "Medicaid",              "spending"),
    ("Table 3-1", 14, "Other mandatory",       "spending"),
    ("Table 3-1", 15, "Offsetting receipts",   "spending"),
    ("Table 3-1", 19, "Defense",               "spending"),
    ("Table 3-1", 20, "Nondefense",            "spending"),
    ("Table 3-1", 22, "Net interest",          "spending"),
    # Revenue by category (Table 1-1)
    ("Table 1-1", 12, "Individual income tax", "revenue"),
    ("Table 1-1", 13, "Payroll tax",           "revenue"),
    ("Table 1-1", 14, "Corporate income tax",  "revenue"),
    ("Table 1-1", 15, "Customs duties",        "revenue"),
    ("Table 1-1", 16, "Other revenue",         "revenue"),
    # Macroeconomic anchors
    ("Table 1-1", 33, "GDP",                   "gdp"),
    ("Table 1-1", 31, "Debt held by public",   "debt"),
    # Avg interest rate on debt held by the public (Table 1-3)
    ("Table 1-3", 37, "Average interest rate", "rate"),
]

# Single-cell scalars from the same workbook:
#   end-FY2025 debt = Table 1-1 row 31 col 2 (Actual 2025) — seed for compounding
#   2025 GDP        = Table 1-1 row 33 col 2 (Actual 2025) — reference for display

def main():
    if not INPUT_XLSX.exists():
        print(f"ERROR: input file not found: {INPUT_XLSX}", file=sys.stderr)
        print("Download from https://www.cbo.gov/publication/61882", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)

    out_rows = []
    out_rows.append(["series", "kind"] + [str(y) for y in YEARS])

    for sheet, row, label, kind in ROWS:
        ws = wb[sheet]
        # Columns 3..11 in the XLSX correspond to 2026..2034
        vals = [ws.cell(row, 3 + i).value for i in range(len(YEARS))]
        # Convert percentage rates: CBO shows interest rate as e.g. 3.404 (percent), we want 0.03404
        if kind == "rate":
            vals = [(v / 100.0) if isinstance(v, (int, float)) else v for v in vals]
        # Validate
        for i, v in enumerate(vals):
            if not isinstance(v, (int, float)):
                raise RuntimeError(f"Bad cell at {sheet}!R{row}C{3+i} ({label}, {YEARS[i]}): {v!r}")
        out_rows.append([label, kind] + [f"{v:.4f}" if kind == "rate" else f"{v:.1f}" for v in vals])

    # Scalars — seed start-of-FY2026 debt = end-of-FY2025 debt (CBO Table 1-1 row 31 col 2 = 2025 actual)
    ws11 = wb["Table 1-1"]
    seed_debt = ws11.cell(31, 2).value   # 30172.402
    gdp_2025 = ws11.cell(33, 2).value    # 30362.025
    if not isinstance(seed_debt, (int, float)) or not isinstance(gdp_2025, (int, float)):
        raise RuntimeError("Could not read seed debt / 2025 GDP from Table 1-1")

    # Put these as single-year "seed" rows for symmetry. Empty cells for 2027..2034.
    seed_row = ["Starting debt end-FY2025", "seed", f"{seed_debt:.1f}"] + ["" for _ in YEARS[1:]]
    gdp_2025_row = ["GDP 2025 actual", "seed", f"{gdp_2025:.1f}"] + ["" for _ in YEARS[1:]]
    out_rows.append(seed_row)
    out_rows.append(gdp_2025_row)

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_CSV.open("w", newline="") as f:
        csv.writer(f).writerows(out_rows)

    print(f"Wrote {OUTPUT_CSV}")
    print(f"  {len(out_rows) - 1} data rows, {len(YEARS)} year columns ({YEARS[0]}-{YEARS[-1]})")
    print(f"  Seed: end-FY2025 debt = {seed_debt:.1f} B")
    print(f"  2034 baseline: debt = {[r for r in out_rows if r[0] == 'Debt held by public'][0][-1]} B, "
          f"GDP = {[r for r in out_rows if r[0] == 'GDP'][0][-1]} B")

if __name__ == "__main__":
    main()
